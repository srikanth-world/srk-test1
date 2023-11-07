import openpyxl

# Load your Excel file
workbook = openpyxl.load_workbook('your_excel_file.xlsx')

# Iterate through each sheet in the workbook
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]

    # Initialize variables to keep track of row counts
    old_data_row_count = 0
    new_data_row_count = 0

    # Flag to indicate when to start counting new data
    counting_new_data = False

    # Iterate through rows in the sheet
    for row in sheet.iter_rows():
        if counting_new_data:
            new_data_row_count += 1
        else:
            old_data_row_count += 1

        for cell in row:
            if cell.value == "QA03 - New":
                counting_new_data = True
                break  # No need to search for "QA03 - New" in the same row

    # Subtract 3 from each count to exclude the header rows and the "QA03 - New" row
    old_data_row_count -= 3
    new_data_row_count -= 3

    # Write the row counts at the end of the sheet
    sheet.cell(row=sheet.max_row + 1, column=1, value="Old Data Row Count:")
    sheet.cell(row=sheet.max_row, column=2, value=old_data_row_count)
    sheet.cell(row=sheet.max_row + 2, column=1, value="New Data Row Count:")
    sheet.cell(row=sheet.max_row + 2, column=2, value=new_data_row_count)

# Save the modified workbook
workbook.save('your_updated_excel_file.xlsx')
